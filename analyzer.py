# -*- coding: utf-8 -*-
"""
Created on Mon Mar 23 15:06:10 2020

@author: Dell
"""

# -*- coding: utf-8 -*-
"""
Created on Sat Mar 21 16:30:52 2020

@author: Dell
"""
import math
import pandas as pd
import requests 
import bs4
import os 
import openpyxl
import matplotlib.pyplot as plt

# change working directory to make a new folder in that directory
os.chdir(r"H:\machine learning\Udemy - Automate the Boring Stuff with Python Programming\noodling") 
current_folder = os.path.join(os.getcwd(), "covid_19 data") 
 
def download_resource(data_folder, url):
    import io, zipfile
    
    if not os.path.exists(data_folder): # create folder if none exist
     os.makedirs(current_folder)
     
    r = requests.get(url, stream = True)
    file_name = r"\covid_file.xlsx"
    output = open(current_folder+file_name, "wb")
    output.write(r.content)
    output.close()

file_url = "https://covid.ourworldindata.org/data/owid-covid-data.xlsx"    
download_resource(current_folder,file_url )

print("[INFO] reading file.............")
wb = openpyxl.load_workbook(r"H:\machine learning\Udemy - Automate the Boring Stuff with Python Programming\noodling\covid_19 data\covid_file.xlsx")
sheet = wb.active
no_of_infected = []

for row in range(2,sheet.max_row+1):
    if sheet.cell(row = row , column =1).value == "DEU": # country code can be changed here
        no_of_infected.append(sheet.cell(row = row, column = 5).value)

no_of_infected = [i for i in no_of_infected if i] #remove nonetype values from the dataset 
new_infections_l=[]
percentage_1=[]
growth_factor_1=[]

for people in range(len(no_of_infected)):
    if people == len(no_of_infected) -1: # so that the loop does not go out of index
        break
    else:
        new_infections=no_of_infected[people+1]-no_of_infected[people]
        new_infections_l.append(new_infections)
        percentage=new_infections_l[people]/no_of_infected[people]
        percentage_1.append(percentage*100)
        growth_factor=no_of_infected[people+1]/no_of_infected[people]
        growth_factor_1.append(growth_factor) # the index goes out of range but the main thing to look at is in growth_factor_!
        
days = [i+1 for i in range(len(growth_factor_1))]

    
plt.plot(days,percentage_1)
plt.title("infection stats")
plt.xlabel("days")
plt.ylabel("infections")
plt.yscale("log")
plt.show()


